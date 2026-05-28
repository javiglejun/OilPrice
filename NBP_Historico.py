import os
import requests
import pandas as pd
from datetime import datetime
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -------------------------------------------------------------------
# 1) QUITAR AVISOS DE SEGURIDAD POR verify=False
# -------------------------------------------------------------------
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# -------------------------------------------------------------------
# 2) CONFIGURACIÓN BÁSICA
# -------------------------------------------------------------------
commodity = "NATURAL_GAS_GBP"
nombre_archivo = "Historico_OilPrice.xlsx"

# Leer la API key desde una variable de entorno llamada OILPRICE_API_KEY
api_key_value = os.getenv("OILPRICE_API_KEY")

# Si no existe la variable, el script se para
if not api_key_value:
    raise ValueError("No se ha encontrado la variable de entorno OILPRICE_API_KEY")

# Cabecera de autenticación para la API
headers = {"Authorization": f"Token {api_key_value}"}

# Lista donde guardaremos temporalmente el dato nuevo
datos_recolectados = []

print("Iniciando descarga de datos...")

# -------------------------------------------------------------------
# 3) LLAMADA A LA API
# -------------------------------------------------------------------
url = f"https://api.oilpriceapi.com/v1/commodities/{commodity}"

try:
    response = requests.get(url, headers=headers, verify=False)
    response.raise_for_status()

    # Convertimos la respuesta de la API a JSON
    data = response.json().get("data", {})

    if data:
        codigo = data.get("code")
        precio = data.get("current_price", {}).get("value")
        moneda = data.get("current_price", {}).get("currency")
        actualizacion = data.get("current_price", {}).get("last_updated")

        # Convertimos la fecha/hora completa a solo la fecha
        # Ejemplo: 2026-05-28T18:14:00Z -> 2026-05-28
        fecha_precio = pd.to_datetime(actualizacion, errors="coerce").date() if actualizacion else None

        # Guardamos el dato como una fila
        datos_recolectados.append({
            "Commodity": commodity,
            "Codigo": codigo,
            "Precio": precio,
            "Moneda": moneda,
            "Ultima_Actualizacion": actualizacion,
            "Fecha_Precio": fecha_precio,
            "Fecha_Ejecucion": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

        print(f"Datos obtenidos correctamente para: {commodity}")
    else:
        print("La API no devolvió datos en el nodo 'data'.")

except requests.exceptions.RequestException as e:
    print(f"Error al consultar {commodity}: {e}")

# -------------------------------------------------------------------
# 4) SI HAY DATOS, ACTUALIZAMOS EL HISTÓRICO
# -------------------------------------------------------------------
if datos_recolectados:
    df_nuevo = pd.DataFrame(datos_recolectados)

    # Convertimos columnas a tipo correcto
    df_nuevo["Precio"] = pd.to_numeric(df_nuevo["Precio"], errors="coerce")
    df_nuevo["Ultima_Actualizacion"] = pd.to_datetime(df_nuevo["Ultima_Actualizacion"], errors="coerce", utc=True)
    df_nuevo["Fecha_Precio"] = pd.to_datetime(df_nuevo["Fecha_Precio"], errors="coerce")
    df_nuevo["Fecha_Ejecucion"] = pd.to_datetime(df_nuevo["Fecha_Ejecucion"], errors="coerce")

    # ----------------------------------------------------------------
    # 5) SI EL EXCEL YA EXISTE, LO LEEMOS Y LE AÑADIMOS EL NUEVO DATO
    # ----------------------------------------------------------------
    if os.path.exists(nombre_archivo):
        print("📂 Archivo existente detectado. Añadiendo nuevos datos...")

        df_existente = pd.read_excel(nombre_archivo, engine="openpyxl")

        # Asegurar que existen todas las columnas esperadas
        columnas_esperadas = [
            "Commodity",
            "Codigo",
            "Precio",
            "Moneda",
            "Ultima_Actualizacion",
            "Fecha_Precio",
            "Fecha_Ejecucion"
        ]

        for col in columnas_esperadas:
            if col not in df_existente.columns:
                df_existente[col] = None

        # Reordenar columnas
        df_existente = df_existente[columnas_esperadas]

        # Convertir tipos
        df_existente["Precio"] = pd.to_numeric(df_existente["Precio"], errors="coerce")
        df_existente["Ultima_Actualizacion"] = pd.to_datetime(df_existente["Ultima_Actualizacion"], errors="coerce", utc=True)
        df_existente["Fecha_Precio"] = pd.to_datetime(df_existente["Fecha_Precio"], errors="coerce")
        df_existente["Fecha_Ejecucion"] = pd.to_datetime(df_existente["Fecha_Ejecucion"], errors="coerce")

        # Unir histórico + dato nuevo
        df_final = pd.concat([df_existente, df_nuevo], ignore_index=True)

    else:
        print("Creando nuevo archivo histórico por primera vez...")
        df_final = df_nuevo.copy()

    # ----------------------------------------------------------------
    # 6) QUEDARSE SIEMPRE CON EL ÚLTIMO PRECIO DEL DÍA
    # ----------------------------------------------------------------
    # Ordena las filas desde la más antigua a la más reciente
    df_final = df_final.sort_values(
        by=["Fecha_Precio", "Ultima_Actualizacion", "Fecha_Ejecucion"],
        ascending=[True, True, True]
    )

    # Si hay varias filas del mismo día para la misma commodity,
    # se queda con la última (la más reciente)
    df_final = df_final.drop_duplicates(
        subset=["Commodity", "Fecha_Precio"],
        keep="last"
    )

    # Orden final por fecha
    df_final = df_final.sort_values(by=["Fecha_Precio"]).reset_index(drop=True)

    # ----------------------------------------------------------------
    # 7) GUARDAR EN EXCEL
    # ----------------------------------------------------------------
    df_final.to_excel(nombre_archivo, index=False, engine="openpyxl")

    # ----------------------------------------------------------------
    # 8) DAR FORMATO BONITO AL EXCEL
    # ----------------------------------------------------------------
    try:
        wb = load_workbook(nombre_archivo)
        ws = wb.active
        ws.title = "Histórico Precios"

        # Estilo de cabecera
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")

        # Aplicar estilo a la primera fila (cabecera)
        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Aplicar formatos de número y fecha
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # A = Commodity
            # B = Codigo
            # C = Precio
            # D = Moneda
            # E = Ultima_Actualizacion
            # F = Fecha_Precio
            # G = Fecha_Ejecucion

            row[2].number_format = '#,##0.00'            # Precio
            row[4].number_format = 'yyyy-mm-dd hh:mm:ss' # Ultima_Actualizacion
            row[5].number_format = 'yyyy-mm-dd'          # Fecha_Precio
            row[6].number_format = 'yyyy-mm-dd hh:mm:ss' # Fecha_Ejecucion

        # Ajustar ancho de columnas automáticamente
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        wb.save(nombre_archivo)
        print(f"✅ Datos guardados correctamente en: {nombre_archivo}")

    except Exception as e:
        print(f"⚠️ Alerta al aplicar estilos visuales: {e}")

else:
    print("No se pudieron obtener datos.")

Add script
